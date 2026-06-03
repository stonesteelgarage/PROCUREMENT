import os
import re
import json
import base64
import sqlite3
from datetime import datetime, timedelta

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from rapidfuzz import fuzz
from openai import OpenAI
import requests
from bs4 import BeautifulSoup
from duckduckgo_search import DDGS
from urllib.parse import urlparse, quote_plus


# ======================================================
# CONFIG IBRIDA: STREAMLIT SECRETS + CONFIG.PY LOCALE
# ======================================================

def get_secret(key, default=None):
    try:
        if key in st.secrets:
            return st.secrets[key]
    except Exception:
        pass

    try:
        import config
        if hasattr(config, key):
            return getattr(config, key)
    except Exception:
        pass

    return default


OPENAI_API_KEY = get_secret("OPENAI_API_KEY")
DB_PATH = get_secret("DB_PATH", "procurement_intelligence.db")
MATCH_THRESHOLD = int(get_secret("MATCH_THRESHOLD", 78))
LOGIN_PASSWORD = get_secret("LOGIN_PASSWORD", "1234")
LOGO_PATH = get_secret("LOGO_PATH", "logo.png")
BACKGROUND_PATH = get_secret("BACKGROUND_PATH", "sfondo")


# ======================================================
# CONFIG BASE
# ======================================================

APP_NAME = "PROCUREMENT INTELLIGENCE"
USER_NAME = "Amandorla"

os.makedirs("uploads", exist_ok=True)
os.makedirs("output", exist_ok=True)
os.makedirs("logs", exist_ok=True)

st.set_page_config(
    page_title=APP_NAME,
    layout="wide",
    initial_sidebar_state="expanded"
)

if not OPENAI_API_KEY:
    st.error("OPENAI_API_KEY non trovata. Inseriscila in config.py oppure nei secrets di Streamlit.")
    st.stop()


# ======================================================
# SFONDO ROBUSTO
# ======================================================

def resolve_file_path(path_value, default_stem):
    candidates = []

    if path_value:
        candidates.append(str(path_value))
        root, ext = os.path.splitext(str(path_value))
        if not ext:
            candidates.extend([
                f"{path_value}.png",
                f"{path_value}.PNG",
                f"{path_value}.jpg",
                f"{path_value}.JPG",
                f"{path_value}.jpeg",
                f"{path_value}.JPEG",
                f"{path_value}.webp",
                f"{path_value}.WEBP",
            ])

    candidates.extend([
        default_stem,
        f"{default_stem}.png",
        f"{default_stem}.PNG",
        f"{default_stem}.jpg",
        f"{default_stem}.JPG",
        f"{default_stem}.jpeg",
        f"{default_stem}.JPEG",
        f"{default_stem}.webp",
        f"{default_stem}.WEBP",
    ])

    for candidate in candidates:
        if candidate and os.path.exists(candidate):
            return candidate

    return None


def get_base64_image(path):
    try:
        if path and os.path.exists(path):
            with open(path, "rb") as f:
                return base64.b64encode(f.read()).decode()
    except Exception:
        pass
    return None


resolved_background_path = resolve_file_path(BACKGROUND_PATH, "sfondo")
background_image = get_base64_image(resolved_background_path)

if background_image:
    ext = os.path.splitext(resolved_background_path)[1].lower()
    mime = "image/png"
    if ext in [".jpg", ".jpeg"]:
        mime = "image/jpeg"
    elif ext == ".webp":
        mime = "image/webp"

    background_css = f'''
    background-image:
        linear-gradient(rgba(2, 10, 20, 0.78), rgba(2, 10, 20, 0.78)),
        url("data:{mime};base64,{background_image}");
    background-size: cover;
    background-position: center center;
    background-repeat: no-repeat;
    background-attachment: fixed;
    '''
else:
    background_css = '''
    background: linear-gradient(135deg, #071526 0%, #0b2238 45%, #08111f 100%);
    '''


# ======================================================
# CSS DARK CORPORATE
# ======================================================

st.markdown(f"""
<style>
.stApp {{
    {background_css}
    color: white;
}}

[data-testid="stSidebar"] {{
    background-color: rgba(6, 18, 33, 0.95);
}}

.main .block-container {{
    background-color: rgba(2, 10, 20, 0.18);
    border-radius: 18px;
    padding-top: 2rem;
}}

h1, h2, h3, h4, h5, h6, p, label, span {{
    color: white !important;
}}

.stButton > button,
.stDownloadButton > button,
button[kind="primary"],
button[kind="secondary"] {{
    background-color: #b30000 !important;
    color: white !important;
    border: 1px solid #ff4d4d !important;
    border-radius: 8px !important;
    font-weight: 700 !important;
}}

.stButton > button:hover,
.stDownloadButton > button:hover,
button[kind="primary"]:hover,
button[kind="secondary"]:hover {{
    background-color: #d00000 !important;
    color: white !important;
    border: 1px solid #ff8080 !important;
}}

.stFileUploader button {{
    background-color: #b30000 !important;
    color: white !important;
    border: 1px solid #ff4d4d !important;
    border-radius: 8px !important;
    font-weight: 700 !important;
}}

.stFileUploader button:hover {{
    background-color: #d00000 !important;
    color: white !important;
}}

div[data-testid="stMetric"] {{
    background-color: rgba(8, 33, 61, 0.85);
    border: 1px solid rgba(0, 210, 255, 0.35);
    border-radius: 14px;
    padding: 18px;
}}

.card {{
    background: rgba(8, 33, 61, 0.82);
    border: 1px solid rgba(0, 210, 255, 0.25);
    border-radius: 18px;
    padding: 22px;
    margin-bottom: 18px;
}}

.hero {{
    background: linear-gradient(90deg, rgba(8,33,61,0.95), rgba(0,80,120,0.35));
    border: 1px solid rgba(0,210,255,0.35);
    border-radius: 22px;
    padding: 30px;
    margin-bottom: 25px;
}}

.small-muted {{
    color: #b8c7d9 !important;
    font-size: 14px;
}}

.stAlert {{
    border-radius: 12px !important;
}}

[data-testid="stDataFrame"] {{
    background: rgba(8, 33, 61, 0.70) !important;
    border-radius: 12px !important;
}}

input[type="password"] {{
    background-color: white !important;
    color: black !important;
    -webkit-text-security: disc !important;
}}

input[type="text"],
textarea {{
    background-color: white !important;
    color: black !important;
}}

input[type="password"]::placeholder,
input[type="text"]::placeholder,
textarea::placeholder {{
    color: #333333 !important;
}}

div[data-baseweb="input"] input {{
    background-color: white !important;
    color: black !important;
}}

div[data-baseweb="textarea"] textarea {{
    background-color: white !important;
    color: black !important;
}}

</style>
""", unsafe_allow_html=True)


# ======================================================
# LOGIN
# ======================================================

def login_screen():
    st.markdown("<br><br>", unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 1.2, 1])

    with col2:
        if os.path.exists(LOGO_PATH):
            st.image(LOGO_PATH, width=180)

        st.markdown(f"""
        <div class="card">
            <h2>{APP_NAME}</h2>
            <p class="small-muted">Vendor Intelligence • Market Intelligence • AI Matching</p>
        </div>
        """, unsafe_allow_html=True)

        password = st.text_input("Password", type="password")

        if st.button("Accedi", use_container_width=True):
            if password == LOGIN_PASSWORD:
                st.session_state["logged_in"] = True
                st.rerun()
            else:
                st.error("Password errata")


if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

if not st.session_state["logged_in"]:
    login_screen()
    st.stop()


# ======================================================
# UTILITY TESTO
# ======================================================

def clean_text(value):
    if value is None:
        return ""

    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    value = str(value)
    value = value.replace("\n", " ")
    value = value.replace("\t", " ")
    value = value.strip()
    value = re.sub(r"\s+", " ", value)

    return value


def normalize_for_match(value):
    value = clean_text(value).lower()
    value = value.replace("'", "")
    value = value.replace("’", "")
    value = value.replace(".", "")
    value = value.replace(",", "")
    value = value.replace("à", "a")
    value = value.replace("è", "e")
    value = value.replace("é", "e")
    value = value.replace("ì", "i")
    value = value.replace("ò", "o")
    value = value.replace("ù", "u")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def extract_email_from_text(text):
    text = clean_text(text)
    found = re.findall(
        r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}",
        text
    )
    return "; ".join(dict.fromkeys(found))


def extract_phone_from_text(text):
    text = clean_text(text)
    found = re.findall(
        r"(?:\+39\s?)?(?:0\d{1,4}[\s./-]?\d{5,8}|3\d{2}[\s./-]?\d{6,7})",
        text
    )
    return "; ".join(dict.fromkeys(found))


def extract_vat_from_text(text):
    text = clean_text(text)
    found = re.findall(
        r"(?:P\.?\s?IVA|Partita\s+IVA|VAT)[:\s]*([0-9]{11})",
        text,
        flags=re.IGNORECASE
    )
    return "; ".join(dict.fromkeys(found))


# ======================================================
# DATABASE CON MIGRAZIONE AUTOMATICA
# ======================================================

def get_conn():
    return sqlite3.connect(DB_PATH, check_same_thread=False)


def init_db():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS suppliers (
        id INTEGER PRIMARY KEY AUTOINCREMENT
    )
    """)

    cur.execute("PRAGMA table_info(suppliers)")
    existing_columns = [row[1] for row in cur.fetchall()]

    required_columns = {
        "package_name": "TEXT",
        "supplier_name": "TEXT",
        "vat_number": "TEXT",
        "email": "TEXT",
        "phone": "TEXT",
        "address_nl1": "TEXT",
        "address_nl2": "TEXT",
        "website": "TEXT",
        "source_file": "TEXT",
        "created_at": "TEXT"
    }

    for column_name, column_type in required_columns.items():
        if column_name not in existing_columns:
            cur.execute(f"ALTER TABLE suppliers ADD COLUMN {column_name} {column_type}")

    conn.commit()
    conn.close()


def insert_supplier(row):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    INSERT INTO suppliers (
        package_name,
        supplier_name,
        vat_number,
        email,
        phone,
        address_nl1,
        address_nl2,
        website,
        source_file,
        created_at
    )
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        row.get("package_name", ""),
        row.get("supplier_name", ""),
        row.get("vat_number", ""),
        row.get("email", ""),
        row.get("phone", ""),
        row.get("address_nl1", ""),
        row.get("address_nl2", ""),
        row.get("website", ""),
        row.get("source_file", ""),
        datetime.now().isoformat(timespec="seconds")
    ))

    conn.commit()
    conn.close()


def load_suppliers():
    conn = get_conn()
    df = pd.read_sql_query("SELECT * FROM suppliers", conn)
    conn.close()
    return df


def supplier_exists(supplier_name, vat_number=""):
    memory = load_suppliers()

    if memory.empty:
        return False

    supplier_name = clean_text(supplier_name).lower()
    vat_number = clean_text(vat_number).lower()

    if vat_number and "vat_number" in memory.columns:
        existing_vat = memory["vat_number"].fillna("").astype(str).str.lower().str.strip()
        if vat_number in existing_vat.values:
            return True

    if "supplier_name" in memory.columns:
        existing_names = memory["supplier_name"].fillna("").astype(str).str.lower().str.strip()
        return supplier_name in existing_names.values

    return False


init_db()


# ======================================================
# LETTURA EXCEL ROBUSTA CON OPENPYXL
# ======================================================

def read_excel_raw(file):
    file.seek(0)

    wb = load_workbook(
        file,
        read_only=True,
        data_only=True
    )

    best_sheet_data = None
    best_score = -1
    best_sheet_name = None

    keywords = [
        "risorsa merceologico",
        "merceologico",
        "nome societa",
        "nome società",
        "societa",
        "società",
        "fornitore",
        "telefono",
        "mail",
        "email",
        "referente",
        "piva",
        "partita iva"
    ]

    for ws in wb.worksheets:
        rows = []

        max_row = min(ws.max_row or 1, 2000)
        max_col = min(ws.max_column or 1, 80)

        for row in ws.iter_rows(
            min_row=1,
            max_row=max_row,
            max_col=max_col,
            values_only=True
        ):
            cleaned = [clean_text(v) for v in row]
            rows.append(cleaned)

        score = 0

        for r in rows[:80]:
            joined = " ".join([x.lower() for x in r if x])
            for kw in keywords:
                if kw in joined:
                    score += 1

        non_empty_rows = sum(1 for r in rows if any(clean_text(x) for x in r))
        score += min(non_empty_rows, 50) / 100

        if score > best_score:
            best_score = score
            best_sheet_data = rows
            best_sheet_name = ws.title

    if not best_sheet_data:
        raise Exception("Nessun foglio Excel leggibile trovato.")

    df = pd.DataFrame(best_sheet_data).fillna("")
    df.attrs["sheet_name"] = best_sheet_name

    return df


# ======================================================
# RICONOSCIMENTO COLONNE
# ======================================================

def normalize_col(c):
    return normalize_for_match(c)


def find_col(df, keywords):
    for col in df.columns:
        col_norm = normalize_col(col)

        for kw in keywords:
            if normalize_for_match(kw) in col_norm:
                return col

    return None


def detect_columns(df):
    return {
        "package_col": find_col(df, [
            "pacchetto",
            "lavorazione",
            "descrizione",
            "categoria",
            "scope",
            "risorsa merceologico",
            "merceologico",
            "risorsa"
        ]),
        "supplier_col": find_col(df, [
            "fornitore",
            "supplier",
            "vendor",
            "ragione sociale",
            "azienda",
            "impresa",
            "nome societa",
            "nome società",
            "societa",
            "società",
            "nome"
        ]),
        "vat_col": find_col(df, [
            "piva",
            "p.iva",
            "partita iva",
            "vat"
        ]),
        "email_col": find_col(df, [
            "email",
            "mail",
            "e-mail"
        ]),
        "phone_col": find_col(df, [
            "telefono",
            "phone",
            "tel",
            "cellulare"
        ]),
        "address1_col": find_col(df, [
            "nl1",
            "indirizzo",
            "address",
            "sede"
        ]),
        "address2_col": find_col(df, [
            "nl2",
            "indirizzo 2",
            "address 2"
        ]),
        "website_col": find_col(df, [
            "sito",
            "website",
            "url",
            "web"
        ])
    }


def detect_header_row_and_columns_local(raw_df):
    best = None

    header_keywords = [
        "pacchetto",
        "lavorazione",
        "risorsa merceologico",
        "merceologico",
        "fornitore",
        "nome societa",
        "nome società",
        "societa",
        "società",
        "telefono",
        "referente",
        "mail",
        "email",
        "partita iva",
        "piva"
    ]

    for idx in range(min(len(raw_df), 80)):
        row_values = [clean_text(v).lower() for v in raw_df.iloc[idx].tolist()]
        joined = " | ".join(row_values)

        score = 0

        for kw in header_keywords:
            if kw in joined:
                score += 1

        non_empty = sum(1 for v in row_values if v)
        if non_empty >= 3:
            score += 1

        if best is None or score > best["score"]:
            best = {
                "row_index": idx,
                "score": score,
                "values": row_values
            }

    if not best or best["score"] < 2:
        return None

    header_idx = best["row_index"]
    headers = [clean_text(v) for v in raw_df.iloc[header_idx].tolist()]

    df = raw_df.iloc[header_idx + 1:].copy()
    df.columns = headers
    df = df.loc[:, [c for c in df.columns if str(c).strip() != ""]]
    df = df.fillna("")

    cols = detect_columns(df)

    return {
        "header_row_index": header_idx,
        "df": df,
        "cols": cols
    }


# ======================================================
# OPENAI HEADER DETECTION
# ======================================================

def ask_openai_excel_mapping(raw_df):
    client = OpenAI(api_key=OPENAI_API_KEY)

    preview_rows = []
    max_rows = min(len(raw_df), 40)
    max_cols = min(raw_df.shape[1], 30)

    for i in range(max_rows):
        row = []

        for j in range(max_cols):
            val = clean_text(raw_df.iat[i, j])

            if val:
                row.append({
                    "col_index": j,
                    "value": val[:160]
                })

        if row:
            preview_rows.append({
                "row_index": i,
                "cells": row
            })

    prompt = f"""
Sei un esperto di vendor list procurement in Excel.

Devi analizzare una preview di un file Excel senza intestazioni standard.
Devi capire:
1. qual è la riga delle intestazioni;
2. quale colonna contiene la lavorazione/pacchetto/risorsa merceologica;
3. quale colonna contiene il nome fornitore/società;
4. se presenti, telefono, email, partita IVA, indirizzi, sito web.

La numerazione di righe e colonne parte da 0.

Preview:
{json.dumps(preview_rows, ensure_ascii=False)}

Rispondi SOLO con JSON valido, senza markdown, con questa struttura esatta:
{{
  "header_row_index": 0,
  "package_col_index": null,
  "supplier_col_index": null,
  "vat_col_index": null,
  "email_col_index": null,
  "phone_col_index": null,
  "address1_col_index": null,
  "address2_col_index": null,
  "website_col_index": null,
  "notes": ""
}}
"""

    response = client.responses.create(
        model="gpt-4.1-mini",
        input=prompt
    )

    txt = response.output_text.strip()
    txt = txt.replace("```json", "").replace("```", "").strip()

    try:
        return json.loads(txt)
    except Exception:
        raise Exception(f"OpenAI non ha restituito JSON valido: {txt}")


def dataframe_from_ai_mapping(raw_df, mapping):
    header_idx = mapping.get("header_row_index")

    if header_idx is None:
        raise Exception("OpenAI non ha individuato la riga intestazioni.")

    header_idx = int(header_idx)

    headers = [clean_text(v) for v in raw_df.iloc[header_idx].tolist()]

    df = raw_df.iloc[header_idx + 1:].copy()
    df.columns = headers
    df = df.fillna("")
    df = df.loc[:, [c for c in df.columns if str(c).strip() != ""]]

    def col_name(index):
        if index is None:
            return None

        try:
            index = int(index)
            if index < len(headers):
                return headers[index]
        except Exception:
            return None

        return None

    cols = {
        "package_col": col_name(mapping.get("package_col_index")),
        "supplier_col": col_name(mapping.get("supplier_col_index")),
        "vat_col": col_name(mapping.get("vat_col_index")),
        "email_col": col_name(mapping.get("email_col_index")),
        "phone_col": col_name(mapping.get("phone_col_index")),
        "address1_col": col_name(mapping.get("address1_col_index")),
        "address2_col": col_name(mapping.get("address2_col_index")),
        "website_col": col_name(mapping.get("website_col_index")),
    }

    return df, cols


def smart_read_vendor_excel(file):
    raw_df = read_excel_raw(file)

    local = detect_header_row_and_columns_local(raw_df)

    if local:
        df = local["df"]
        cols = local["cols"]

        if cols.get("package_col") and cols.get("supplier_col"):
            return df, cols, f"Riconoscimento automatico locale: intestazioni alla riga Excel {local['header_row_index'] + 1}"

    mapping = ask_openai_excel_mapping(raw_df)
    df, cols = dataframe_from_ai_mapping(raw_df, mapping)

    if not cols.get("package_col") or not cols.get("supplier_col"):
        raise Exception(f"OpenAI non ha individuato pacchetto e fornitore. Dettagli: {mapping}")

    return df, cols, f"Riconoscimento OpenAI: intestazioni alla riga Excel {int(mapping.get('header_row_index')) + 1}"


def smart_read_template_excel(file):
    raw_df = read_excel_raw(file)

    local = detect_header_row_and_columns_local(raw_df)

    if local:
        df = local["df"]
        cols = local["cols"]

        if cols.get("package_col"):
            return df, cols, f"Riconoscimento automatico locale: intestazioni alla riga Excel {local['header_row_index'] + 1}"

    mapping = ask_openai_excel_mapping(raw_df)
    df, cols = dataframe_from_ai_mapping(raw_df, mapping)

    if not cols.get("package_col"):
        raise Exception(f"OpenAI non ha individuato la colonna pacchetto. Dettagli: {mapping}")

    return df, cols, f"Riconoscimento OpenAI: intestazioni alla riga Excel {int(mapping.get('header_row_index')) + 1}"


# ======================================================
# IMPORT VENDOR
# ======================================================

def import_vendor_excel(file):
    df, cols, detection_msg = smart_read_vendor_excel(file)

    count = 0
    skipped = 0

    for _, r in df.iterrows():
        package = clean_text(r.get(cols["package_col"], ""))
        supplier = clean_text(r.get(cols["supplier_col"], ""))

        if not package or not supplier:
            skipped += 1
            continue

        if supplier.lower() in [
            "nome societa'",
            "nome società",
            "nome societa",
            "fornitore",
            "supplier",
            "vendor",
            "azienda",
            "impresa"
        ]:
            skipped += 1
            continue

        email = clean_text(r.get(cols["email_col"], "")) if cols.get("email_col") else ""
        phone = clean_text(r.get(cols["phone_col"], "")) if cols.get("phone_col") else ""
        vat_number = clean_text(r.get(cols["vat_col"], "")) if cols.get("vat_col") else ""

        row_text = " ".join([clean_text(v) for v in r.tolist()])

        if not email:
            email = extract_email_from_text(row_text)

        if not phone:
            phone = extract_phone_from_text(row_text)

        if not vat_number:
            vat_number = extract_vat_from_text(row_text)

        insert_supplier({
            "package_name": package,
            "supplier_name": supplier,
            "vat_number": vat_number,
            "email": email,
            "phone": phone,
            "address_nl1": clean_text(r.get(cols["address1_col"], "")) if cols.get("address1_col") else "",
            "address_nl2": clean_text(r.get(cols["address2_col"], "")) if cols.get("address2_col") else "",
            "website": clean_text(r.get(cols["website_col"], "")) if cols.get("website_col") else "",
            "source_file": file.name
        })

        count += 1

    return count, detection_msg, skipped


# ======================================================
# MATCHING E GENERAZIONE VENDOR
# ======================================================

def match_package(package, memory):
    results = []

    if memory.empty:
        return results

    for _, row in memory.iterrows():
        score = fuzz.token_set_ratio(
            str(package).lower(),
            str(row["package_name"]).lower()
        )

        if score >= MATCH_THRESHOLD:
            item = row.to_dict()
            item["score"] = score
            results.append(item)

    return sorted(results, key=lambda x: x["score"], reverse=True)


def generate_vendor_from_template(template_file):
    template_df, cols, detection_msg = smart_read_template_excel(template_file)

    if not cols["package_col"]:
        raise Exception("Nel template non trovo la colonna pacchetto/lavorazione.")

    memory = load_suppliers()

    if memory.empty:
        raise Exception("Database storico vuoto. Importa prima almeno una vendor.")

    output_rows = []
    preview = []

    for _, row in template_df.iterrows():
        package = clean_text(row.get(cols["package_col"], ""))

        if not package:
            continue

        matches = match_package(package, memory)

        seen = set()
        clean_matches = []

        for m in matches:
            key = str(m.get("vat_number") or m.get("supplier_name")).lower().strip()

            if key and key not in seen:
                seen.add(key)
                clean_matches.append(m)

        preview.append({
            "Pacchetto": package,
            "Fornitori trovati": len(clean_matches),
            "Miglior match": clean_matches[0]["package_name"] if clean_matches else "",
            "Score": clean_matches[0]["score"] if clean_matches else ""
        })

        if clean_matches:
            for supplier in clean_matches:
                new_row = row.copy()

                if cols["supplier_col"]:
                    new_row[cols["supplier_col"]] = supplier.get("supplier_name", "")
                if cols["vat_col"]:
                    new_row[cols["vat_col"]] = supplier.get("vat_number", "")
                if cols["email_col"]:
                    new_row[cols["email_col"]] = supplier.get("email", "")
                if cols["phone_col"]:
                    new_row[cols["phone_col"]] = supplier.get("phone", "")
                if cols["address1_col"]:
                    new_row[cols["address1_col"]] = supplier.get("address_nl1", "")
                if cols["address2_col"]:
                    new_row[cols["address2_col"]] = supplier.get("address_nl2", "")
                if cols["website_col"]:
                    new_row[cols["website_col"]] = supplier.get("website", "")

                output_rows.append(new_row)
        else:
            output_rows.append(row)

    output_df = pd.DataFrame(output_rows)
    preview_df = pd.DataFrame(preview)

    output_path = os.path.join(
        "output",
        f"vendor_compilata_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    output_df.to_excel(output_path, index=False)

    return output_path, preview_df, detection_msg



# ======================================================
# SCOUTING WEB REALE: RICERCA + SCRAPING + AI + DB
# ======================================================

def domain_to_supplier_name(url):
    try:
        domain = urlparse(url).netloc.replace("www.", "")
        name = domain.split(".")[0]
        return name.upper()
    except Exception:
        return ""


def search_web_suppliers(package, area="", max_results=10):
    """
    Cerca fornitori sul web partendo dalla lavorazione/package.
    Non richiede URL all'utente.

    Versione robusta:
    1) prova DuckDuckGo tramite libreria DDGS;
    2) se DDGS restituisce 0 risultati, usa fallback HTML su DuckDuckGo;
    3) se ancora 0, usa fallback HTML su Bing.

    In questo modo non rimane più bloccato con "Risultati web analizzati: 0".
    """
    package = clean_text(package)
    area = clean_text(area)

    base_terms = f'{package} {area} Italia'.strip()

    queries = [
        f'azienda fornitore {base_terms} contatti email telefono partita iva',
        f'impresa {base_terms} servizi contatti',
        f'azienda {base_terms} lavori forniture subappalto',
        f'{base_terms} site:.it contatti',
    ]

    results = []
    seen_urls = set()

    def add_result(title, url, snippet, query, source):
        url = clean_text(url)
        title = clean_text(title)
        snippet = clean_text(snippet)

        if not url:
            return

        # scarta risultati chiaramente non utili
        bad_domains = [
            "facebook.com", "instagram.com", "linkedin.com", "youtube.com",
            "amazon.", "wikipedia.org", "subito.it", "ebay.",
            "paginegialle.it/", "paginebianche.it/"
        ]
        url_l = url.lower()
        if any(bad in url_l for bad in bad_domains):
            return

        if url in seen_urls:
            return

        seen_urls.add(url)
        results.append({
            "title": title,
            "url": url,
            "snippet": snippet,
            "query": query,
            "search_source": source,
        })

    # 1) DuckDuckGo tramite libreria
    try:
        with DDGS() as ddgs:
            for query in queries:
                if len(results) >= max_results:
                    break

                try:
                    # Alcune versioni accettano region/safesearch/timelimit/backend, altre no.
                    raw_results = ddgs.text(
                        query,
                        region="it-it",
                        safesearch="moderate",
                        max_results=max_results
                    )
                except TypeError:
                    raw_results = ddgs.text(query, max_results=max_results)

                for r in raw_results or []:
                    add_result(
                        r.get("title", ""),
                        r.get("href", ""),
                        r.get("body", ""),
                        query,
                        "DDGS"
                    )
                    if len(results) >= max_results:
                        break
    except Exception as e:
        st.warning(f"DuckDuckGo libreria non disponibile o senza risultati: {e}")

    # 2) Fallback HTML DuckDuckGo
    if len(results) == 0:
        for query in queries:
            if len(results) >= max_results:
                break
            try:
                search_url = "https://html.duckduckgo.com/html/?q=" + quote_plus(query)
                headers = {
                    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/125 Safari/537.36",
                    "Accept-Language": "it-IT,it;q=0.9,en;q=0.8",
                }
                response = requests.get(search_url, headers=headers, timeout=15)
                soup = BeautifulSoup(response.text, "html.parser")

                for item in soup.select(".result"):
                    a = item.select_one("a.result__a")
                    snippet_tag = item.select_one(".result__snippet")
                    if not a:
                        continue
                    title = a.get_text(" ", strip=True)
                    url = a.get("href", "")
                    snippet = snippet_tag.get_text(" ", strip=True) if snippet_tag else ""
                    add_result(title, url, snippet, query, "DuckDuckGo HTML")
                    if len(results) >= max_results:
                        break
            except Exception as e:
                st.warning(f"Fallback DuckDuckGo HTML non riuscito: {e}")

    # 3) Fallback Bing HTML
    if len(results) == 0:
        for query in queries:
            if len(results) >= max_results:
                break
            try:
                search_url = "https://www.bing.com/search?q=" + quote_plus(query)
                headers = {
                    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/125 Safari/537.36",
                    "Accept-Language": "it-IT,it;q=0.9,en;q=0.8",
                }
                response = requests.get(search_url, headers=headers, timeout=15)
                soup = BeautifulSoup(response.text, "html.parser")

                for item in soup.select("li.b_algo"):
                    a = item.select_one("h2 a")
                    snippet_tag = item.select_one("p")
                    if not a:
                        continue
                    title = a.get_text(" ", strip=True)
                    url = a.get("href", "")
                    snippet = snippet_tag.get_text(" ", strip=True) if snippet_tag else ""
                    add_result(title, url, snippet, query, "Bing HTML")
                    if len(results) >= max_results:
                        break
            except Exception as e:
                st.warning(f"Fallback Bing HTML non riuscito: {e}")

    return results

def scrape_website(url):
    """
    Legge il testo visibile del sito. Se il sito blocca requests, ritorna stringa vuota.
    """
    try:
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0 Safari/537.36"
            ),
            "Accept-Language": "it-IT,it;q=0.9,en;q=0.8",
        }

        response = requests.get(url, headers=headers, timeout=12, allow_redirects=True)

        if response.status_code >= 400:
            return ""

        content_type = response.headers.get("Content-Type", "").lower()
        if "text/html" not in content_type and "application/xhtml" not in content_type:
            return ""

        soup = BeautifulSoup(response.text, "html.parser")

        for tag in soup(["script", "style", "noscript", "svg", "canvas", "iframe"]):
            tag.extract()

        text = soup.get_text(separator=" ")
        text = re.sub(r"\s+", " ", text)
        return clean_text(text[:18000])

    except Exception:
        return ""


def safe_json_loads(txt):
    txt = clean_text(txt)
    txt = txt.replace("```json", "").replace("```", "").strip()

    try:
        return json.loads(txt)
    except Exception:
        pass

    match = re.search(r"\{.*\}", txt, flags=re.DOTALL)
    if match:
        return json.loads(match.group(0))

    raise Exception(f"JSON non valido: {txt[:500]}")


def analyze_web_supplier_with_ai(package, url, title, snippet, page_text):
    client = OpenAI(api_key=OPENAI_API_KEY)

    fallback_name = domain_to_supplier_name(url)
    extracted_email = extract_email_from_text(page_text)
    extracted_phone = extract_phone_from_text(page_text)
    extracted_vat = extract_vat_from_text(page_text)

    prompt = f"""
Sei un esperto procurement per infrastrutture, costruzioni, impiantistica, cantieri, subappalti e forniture tecniche.

Devi valutare se il sito trovato online appartiene a un possibile fornitore pertinente per questa lavorazione/package.

LAVORAZIONE / PACKAGE RICHIESTA:
{package}

URL:
{url}

NOME DA DOMINIO SE NON TROVI ALTRO:
{fallback_name}

TITOLO RISULTATO:
{title}

SNIPPET MOTORE DI RICERCA:
{snippet}

DATI ESTRATTI CON REGEX DAL SITO:
email: {extracted_email}
telefono: {extracted_phone}
partita iva: {extracted_vat}

TESTO LETTO DAL SITO:
{page_text[:10000]}

Restituisci SOLO JSON valido, senza markdown, con questa struttura esatta:

{{
  "is_supplier": true,
  "supplier_name": "",
  "package_name": "{package}",
  "vat_number": "",
  "email": "",
  "phone": "",
  "address_nl1": "",
  "address_nl2": "",
  "website": "{url}",
  "services": "",
  "relevance_score": 0,
  "reason": ""
}}

Regole:
- is_supplier deve essere true solo se sembra una vera azienda/impresa/fornitore, non un portale generico, marketplace, directory pura o articolo.
- relevance_score deve essere da 0 a 100.
- Se l'azienda è pertinente in modo concreto alla lavorazione, assegna almeno 70.
- Se è solo vagamente collegata, assegna 40-60.
- Se non trovi il nome azienda, usa il nome da dominio.
- Se email, telefono o P.IVA non sono presenti, lascia stringa vuota.
- package_name deve restare la lavorazione richiesta dall'utente.
"""

    response = client.responses.create(
        model="gpt-4.1-mini",
        input=prompt
    )

    return safe_json_loads(response.output_text)


def scouting_web_and_save(package, area="", max_results=10, min_score=60):
    """
    Cerca fornitori online, entra nei siti, estrae dati, valuta con AI e salva nel DB SQLite.
    """
    package = clean_text(package)
    area = clean_text(area)

    web_results = search_web_suppliers(package, area, max_results=max_results)

    saved = []
    discarded = []

    for result in web_results:
        url = result.get("url", "")
        title = result.get("title", "")
        snippet = result.get("snippet", "")

        page_text = scrape_website(url)
        if not page_text:
            page_text = snippet

        extracted_email = extract_email_from_text(page_text)
        extracted_phone = extract_phone_from_text(page_text)
        extracted_vat = extract_vat_from_text(page_text)

        try:
            ai_data = analyze_web_supplier_with_ai(
                package=package,
                url=url,
                title=title,
                snippet=snippet,
                page_text=page_text
            )

            supplier_name = clean_text(ai_data.get("supplier_name", "")) or domain_to_supplier_name(url) or title
            vat_number = clean_text(ai_data.get("vat_number", "")) or extracted_vat
            email = clean_text(ai_data.get("email", "")) or extracted_email
            phone = clean_text(ai_data.get("phone", "")) or extracted_phone
            relevance = int(ai_data.get("relevance_score", 0) or 0)
            is_supplier = bool(ai_data.get("is_supplier", False))

            row = {
                "package_name": package,
                "supplier_name": supplier_name,
                "vat_number": vat_number,
                "email": email,
                "phone": phone,
                "address_nl1": clean_text(ai_data.get("address_nl1", "")),
                "address_nl2": clean_text(ai_data.get("address_nl2", "")),
                "website": url,
                "source_file": "WEB_SCOUTING",
            }

            result_row = dict(row)
            result_row["relevance_score"] = relevance
            result_row["reason"] = clean_text(ai_data.get("reason", ""))
            result_row["services"] = clean_text(ai_data.get("services", ""))

            if is_supplier and relevance >= min_score:
                if not supplier_exists(supplier_name, vat_number):
                    insert_supplier(row)
                    saved.append(result_row)
                else:
                    result_row["reason"] = "Già presente nel database"
                    discarded.append(result_row)
            else:
                if not result_row["reason"]:
                    result_row["reason"] = "Non pertinente o sotto soglia"
                discarded.append(result_row)

        except Exception as e:
            discarded.append({
                "package_name": package,
                "supplier_name": title or domain_to_supplier_name(url),
                "website": url,
                "relevance_score": 0,
                "reason": f"Errore analisi: {e}"
            })

    return saved, discarded, web_results


# ======================================================
# SIDEBAR CON TASTI
# ======================================================

with st.sidebar:
    if os.path.exists(LOGO_PATH):
        st.image(LOGO_PATH, width=160)

    st.markdown(f"### {APP_NAME}")
    st.caption("Procurement dashboard")
    st.divider()

    if "section" not in st.session_state:
        st.session_state["section"] = "Dashboard"

    if st.button("Dashboard", use_container_width=True):
        st.session_state["section"] = "Dashboard"

    if st.button("Import Vendor Storiche", use_container_width=True):
        st.session_state["section"] = "Import Vendor Storiche"

    if st.button("Genera Vendor Gara", use_container_width=True):
        st.session_state["section"] = "Genera Vendor Gara"

    if st.button("Scouting", use_container_width=True):
        st.session_state["section"] = "Scouting"

    if st.button("Database", use_container_width=True):
        st.session_state["section"] = "Database"

    section = st.session_state["section"]

    st.divider()
    st.markdown(f"Utente: **{USER_NAME}**")

    if st.button("Logout", use_container_width=True):
        st.session_state["logged_in"] = False
        st.rerun()


# ======================================================
# HEADER
# ======================================================

st.markdown(f"""
<div class="hero">
    <h1>{APP_NAME}</h1>
    <p class="small-muted">Vendor Intelligence • Market Intelligence • AI Matching</p>
</div>
""", unsafe_allow_html=True)


# ======================================================
# DASHBOARD
# ======================================================

memory = load_suppliers()

if section == "Dashboard":
    col1, col2 = st.columns(2)

    total_suppliers = memory["supplier_name"].nunique() if not memory.empty and "supplier_name" in memory.columns else 0

    if not memory.empty and "created_at" in memory.columns:
        memory["created_at_dt"] = pd.to_datetime(memory["created_at"], errors="coerce")
        last_30 = memory[
            memory["created_at_dt"] >= datetime.now() - timedelta(days=30)
        ]["supplier_name"].nunique()
    else:
        last_30 = 0

    with col1:
        st.metric("Fornitori totali nel database", total_suppliers)

    with col2:
        st.metric("Fornitori aggiunti ultimi 30 giorni", last_30)

    st.markdown("""
    <div class="card">
        <h3>Flusso operativo</h3>
        <p>1. Importa vendor storiche Excel</p>
        <p>2. Genera vendor list per nuova gara</p>
        <p>3. Usa scouting per cercare nuovi fornitori</p>
        <p>4. Alimenta progressivamente la memoria storica locale</p>
    </div>
    """, unsafe_allow_html=True)


# ======================================================
# IMPORT VENDOR
# ======================================================

elif section == "Import Vendor Storiche":
    st.header("Import Vendor Storiche")

    files = st.file_uploader(
        "Carica una o più vendor list Excel compilate",
        type=["xlsx"],
        accept_multiple_files=True
    )

    if st.button("Importa nel database", use_container_width=True):
        if not files:
            st.warning("Carica almeno un file Excel.")
        else:
            total = 0

            with st.spinner("ALMOND Intelligence sta lavorando..."):
                for file in files:
                    try:
                        count, msg, skipped = import_vendor_excel(file)
                        total += count

                        st.success("Vendor importata correttamente")
                        st.info(f"Numero fornitori importati nel DB: {count}")
                        st.success("Aggiornamento DB terminato")
                        st.caption(msg)

                    except Exception as e:
                        st.error(f"Errore su {file.name}: {e}")

            st.info(f"Totale fornitori importati nel DB: {total}")


# ======================================================
# GENERA VENDOR
# ======================================================

elif section == "Genera Vendor Gara":
    st.header("Genera Vendor Gara")

    template = st.file_uploader(
        "Carica template vendor list della gara",
        type=["xlsx"]
    )

    if st.button("Genera vendor compilata", use_container_width=True):
        if not template:
            st.warning("Carica prima il template.")
        else:
            try:
                output_path, preview_df, msg = generate_vendor_from_template(template)

                st.caption(msg)

                st.subheader("Anteprima matching")
                st.dataframe(preview_df, use_container_width=True)

                with open(output_path, "rb") as f:
                    st.download_button(
                        "Scarica Excel compilato",
                        data=f,
                        file_name=os.path.basename(output_path),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

                st.success("Vendor list generata correttamente.")

            except Exception as e:
                st.error(str(e))


# ======================================================
# SCOUTING
# ======================================================

elif section == "Scouting":
    st.header("Scouting Web Fornitori")

    st.info(
        "Inserisci una lavorazione: il sistema cerca fornitori nel web, legge i siti, "
        "estrae dati aziendali, valuta la pertinenza con ALMOND AI e aggiorna il database."
    )

    package = st.text_input(
        "Lavorazione / package",
        placeholder="Esempio: carpenteria metallica, impianti antincendio, impermeabilizzazioni gallerie"
    )

    area = st.text_input(
        "Area geografica opzionale",
        placeholder="Esempio: Lazio, Roma, Italia, Nord Italia"
    )

    col_a, col_b = st.columns(2)

    with col_a:
        max_results = st.slider("Risultati web da analizzare", 2, 30, 10)

    with col_b:
        min_score = st.slider("Pertinenza minima per salvare nel DB", 0, 100, 60)

    if st.button("Cerca fornitori nel web e aggiorna DB", use_container_width=True):
        if not package:
            st.warning("Inserisci prima una lavorazione.")
        else:
            with st.spinner("ALMOND Intelligence sta cercando fornitori nel web e leggendo i siti..."):
                saved, discarded, web_results = scouting_web_and_save(
                    package=package,
                    area=area,
                    max_results=max_results,
                    min_score=min_score
                )

            st.success(f"Fornitori nuovi salvati nel DB: {len(saved)}")
            st.info(f"Risultati web trovati: {len(web_results)} | Scartati / già presenti / sotto soglia: {len(discarded)}")

            if web_results:
                with st.expander("Mostra risultati web grezzi trovati"):
                    st.dataframe(pd.DataFrame(web_results), use_container_width=True)
            else:
                st.error("La ricerca web non ha restituito risultati. Questa versione usa fallback HTML DuckDuckGo + Bing; se resta a zero, verifica connessione Internet di Streamlit/Mac o aggiorna duckduckgo-search.")

            if saved:
                st.subheader("Nuovi fornitori salvati nel database")
                df_saved = pd.DataFrame(saved)
                st.dataframe(df_saved, use_container_width=True)

                csv_saved = df_saved.to_csv(index=False).encode("utf-8")
                st.download_button(
                    "Scarica nuovi fornitori CSV",
                    data=csv_saved,
                    file_name="scouting_nuovi_fornitori.csv",
                    mime="text/csv",
                    use_container_width=True
                )

            if discarded:
                st.subheader("Risultati scartati / già presenti / non pertinenti")
                df_discarded = pd.DataFrame(discarded)
                st.dataframe(df_discarded, use_container_width=True)

            if web_results and not saved and not discarded:
                st.warning("Sono stati trovati risultati web, ma nessuno è stato salvato. Prova ad abbassare la pertinenza minima a 40 oppure usa una lavorazione meno specifica.")

    st.divider()

    st.subheader("Ricerca nel database storico")
    st.caption("Questa parte usa i fornitori già presenti nel DB, separata dallo scouting web.")

    db_package = st.text_input(
        "Cerca nel DB storico",
        placeholder="Esempio: impianti elettrici, carpenteria, segnalamento"
    )

    if st.button("Cerca nel database", use_container_width=True):
        if not db_package:
            st.warning("Inserisci una lavorazione.")
        else:
            results = match_package(db_package, load_suppliers())

            if results:
                st.success(f"Trovati {len(results)} fornitori nel database.")
                st.dataframe(pd.DataFrame(results), use_container_width=True)
            else:
                st.warning("Nessun fornitore trovato nel database storico.")


# ======================================================
# DATABASE
# ======================================================

elif section == "Database":
    st.header("Database storico vendor")

    st.metric("Righe totali", len(memory))

    if not memory.empty:
        st.dataframe(memory, use_container_width=True)

        csv = memory.to_csv(index=False).encode("utf-8")

        st.download_button(
            "Scarica database CSV",
            data=csv,
            file_name="database_vendor.csv",
            mime="text/csv",
            use_container_width=True
        )
    else:
        st.info("Database ancora vuoto.")

